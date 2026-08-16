"""Exports Excel et 4W.

Deux regles non negociables du cahier des charges :
  - DCP-03 : la pseudonymisation s'applique aux exports comme aux ecrans ;
  - SEC-09 : tout export de donnees personnelles est journalise.
"""
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView

from activities.models import Activity
from audit.models import AuditLog
from audit.services import journaliser
from authorization.permissions import PermissionMetier
from authorization.scopes import projets_accessibles
from beneficiaries.models import Household, Person
from monitoring.models import Indicator
from referentials.models import Organization, Zone

ENTETE_FOND = PatternFill("solid", fgColor="004429")
ENTETE_POLICE = Font(color="FFFFFF", bold=True)


def _ancetre(zone, niveau):
    while zone is not None and zone.level != niveau:
        zone = zone.parent
    return zone


def _entete(feuille, colonnes):
    feuille.append(colonnes)
    for indice in range(1, len(colonnes) + 1):
        cellule = feuille.cell(row=1, column=indice)
        cellule.fill = ENTETE_FOND
        cellule.font = ENTETE_POLICE
        cellule.alignment = Alignment(vertical="center")
    feuille.freeze_panes = "A2"


def _ajuster(feuille):
    for colonne in feuille.columns:
        largeur = max(
            (len(str(c.value)) for c in colonne if c.value is not None), default=8
        )
        lettre = get_column_letter(colonne[0].column)
        feuille.column_dimensions[lettre].width = min(largeur + 3, 55)


def _reponse(classeur, nom):
    flux = BytesIO()
    classeur.save(flux)
    flux.seek(0)
    reponse = HttpResponse(
        flux.read(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    horodatage = timezone.localdate().isoformat()
    reponse["Content-Disposition"] = (
        f'attachment; filename="kaya_{nom}_{horodatage}.xlsx"'
    )
    return reponse


class ExportExcel(APIView):
    """Base commune : permission, construction du classeur, journalisation."""

    permission_classes = [IsAuthenticated, PermissionMetier]
    permission_codes = {"default": "export.realiser"}
    nom_fichier = "export"

    def remplir(self, classeur, request):  # pragma: no cover - redefini
        raise NotImplementedError

    def get(self, request):
        classeur = Workbook()
        classeur.remove(classeur.active)
        lignes = self.remplir(classeur, request)

        journaliser(
            AuditLog.Action.EXPORT,
            actor=request.user,
            object_type=self.nom_fichier,
            detail=f"Export Excel « {self.nom_fichier} » — {lignes} lignes",
            request=request,
        )
        return _reponse(classeur, self.nom_fichier)


class BeneficiairesExportView(ExportExcel):
    """Menages et individus. Les champs nominatifs suivent l'habilitation."""

    permission_codes = {"default": "beneficiaire.exporter"}
    nom_fichier = "beneficiaires"

    def remplir(self, classeur, request):
        nominatif = request.user.has_permission("beneficiaire.voir_donnees_nominatives")

        menages = Household.objects.filter(
            project_links__project__in=projets_accessibles(request.user)
        ).exclude(
            validation_status=Household.ValidationStatus.DOUBLON
        ).select_related("zone__parent__parent__parent", "registered_by").distinct()

        feuille = classeur.create_sheet("Menages")
        _entete(feuille, [
            "Code menage", "Chef de menage", "Taille", "Region", "Prefecture",
            "Localite", "Statut de residence", "Score de vulnerabilite",
            "Statut de validation", "Enregistre le", "Enregistre par",
        ])

        total = 0
        for menage in menages:
            region = _ancetre(menage.zone, Zone.Level.REGION)
            prefecture = _ancetre(menage.zone, Zone.Level.PREFECTURE)
            feuille.append([
                menage.code,
                menage.head_name if nominatif else menage.code,
                menage.size,
                region.name if region else "",
                prefecture.name if prefecture else "",
                menage.zone.name,
                menage.get_residence_status_display(),
                menage.vulnerability_score,
                menage.get_validation_status_display(),
                menage.registered_at.date() if menage.registered_at else "",
                menage.registered_by.full_name if menage.registered_by else "",
            ])
            total += 1
        _ajuster(feuille)

        # ---- individus, desagreges
        individus = Person.objects.filter(household__in=menages).select_related("household")
        feuille2 = classeur.create_sheet("Individus")
        _entete(feuille2, [
            "Code menage", "Prenom", "Nom", "Sexe", "Age", "Tranche d'age",
            "Lien avec le chef", "Chef de menage", "Scolarise", "Handicap",
        ])

        for personne in individus:
            age = personne.age
            tranche = (
                "" if age is None
                else "0-5" if age <= 5
                else "6-17" if age <= 17
                else "18-59" if age <= 59
                else "60+"
            )
            feuille2.append([
                personne.household.code,
                personne.first_name if nominatif else "—",
                personne.last_name if nominatif else "—",
                personne.get_sex_display(),
                age,
                tranche,
                personne.relation_to_head,
                "oui" if personne.is_head else "non",
                "oui" if personne.is_enrolled else "non",
                "oui" if personne.has_disability else "non",
            ])
            total += 1
        _ajuster(feuille2)

        # ---- note de protection, visible dans le fichier remis
        note = classeur.create_sheet("Mentions")
        note.append(["Export genere par la plateforme Kaya"])
        note.append(["Organisation", str(Organization.get_solo())])
        note.append(["Date", timezone.localdate().isoformat()])
        note.append(["Demandeur", request.user.full_name])
        note.append([
            "Donnees nominatives",
            "incluses" if nominatif else "pseudonymisees : le demandeur n'est pas habilite",
        ])
        note.append(["Perimetre", "limite aux projets accessibles au demandeur"])
        _ajuster(note)

        return total


class ActivitesExportView(ExportExcel):
    nom_fichier = "activites"

    def remplir(self, classeur, request):
        activites = Activity.objects.filter(
            project__in=projets_accessibles(request.user)
        ).select_related("project", "zone__parent__parent__parent", "agent", "validated_by")

        feuille = classeur.create_sheet("Activites")
        _entete(feuille, [
            "Code", "Projet", "Type", "Date de realisation", "Region", "Prefecture",
            "Localite", "Statut", "Hommes", "Femmes", "Total participants",
            "Agent", "Valide par", "Pieces jointes", "Alertes qualite",
        ])

        total = 0
        for activite in activites:
            region = _ancetre(activite.zone, Zone.Level.REGION)
            prefecture = _ancetre(activite.zone, Zone.Level.PREFECTURE)
            effectifs = activite.participants_totaux
            feuille.append([
                activite.code,
                activite.project.code,
                activite.get_type_display(),
                activite.activity_date,
                region.name if region else "",
                prefecture.name if prefecture else "",
                activite.zone.name,
                activite.get_status_display(),
                effectifs.get("hommes", 0),
                effectifs.get("femmes", 0),
                effectifs.get("total", 0),
                activite.agent.full_name,
                activite.validated_by.full_name if activite.validated_by else "",
                activite.attachments.count(),
                " | ".join(activite.alertes_qualite),
            ])
            total += 1

        _ajuster(feuille)
        return total


class QuatreWExportView(ExportExcel):
    """Format 4W des clusters humanitaires : Who, What, Where, When.

    Seules les activites validees y figurent : un 4W transmis a un cluster
    engage l'organisation.
    """

    nom_fichier = "4w"

    def remplir(self, classeur, request):
        organisation = Organization.get_solo()
        activites = Activity.objects.filter(
            project__in=projets_accessibles(request.user),
            status=Activity.Status.VALIDEE,
        ).select_related("project", "zone__parent__parent__parent").prefetch_related(
            "project__sectors"
        )

        feuille = classeur.create_sheet("4W")
        _entete(feuille, [
            "Organisation",              # Who
            "Secteur", "Activite",       # What
            "Region", "Prefecture", "Localite",   # Where
            "Date",                      # When
            "Hommes", "Femmes", "Total beneficiaires",
            "Projet", "Statut du projet",
        ])

        total = 0
        for activite in activites:
            region = _ancetre(activite.zone, Zone.Level.REGION)
            prefecture = _ancetre(activite.zone, Zone.Level.PREFECTURE)
            effectifs = activite.participants_totaux
            secteurs = ", ".join(s.label for s in activite.project.sectors.all())
            feuille.append([
                organisation.acronym or organisation.name,
                secteurs,
                activite.get_type_display(),
                region.name if region else "",
                prefecture.name if prefecture else "",
                activite.zone.name,
                activite.activity_date,
                effectifs.get("hommes", 0),
                effectifs.get("femmes", 0),
                effectifs.get("total", 0),
                activite.project.code,
                activite.project.get_status_display(),
            ])
            total += 1

        _ajuster(feuille)
        return total


class IndicateursExportView(ExportExcel):
    nom_fichier = "indicateurs"

    def remplir(self, classeur, request):
        indicateurs = Indicator.objects.filter(
            element__project__in=projets_accessibles(request.user)
        ).select_related("element__project", "owner").prefetch_related(
            "readings", "disaggregations"
        )

        feuille = classeur.create_sheet("Indicateurs")
        _entete(feuille, [
            "Code", "Projet", "Element du cadre logique", "Intitule", "Unite",
            "Reference", "Cible", "Valeur atteinte", "Taux d'atteinte (%)",
            "Taux attendu (%)", "Statut", "Periodicite", "Mode de calcul",
            "Desagregations", "Responsable",
        ])

        total = 0
        for indicateur in indicateurs:
            feuille.append([
                indicateur.code,
                indicateur.element.project.code,
                indicateur.element.code,
                indicateur.title,
                indicateur.get_unit_display(),
                float(indicateur.baseline),
                float(indicateur.target),
                float(indicateur.valeur_atteinte),
                indicateur.taux_atteinte,
                indicateur.taux_attendu,
                indicateur.statut_atteinte,
                indicateur.get_frequency_display(),
                indicateur.get_computation_mode_display(),
                ", ".join(d.get_dimension_display() for d in indicateur.disaggregations.all()),
                indicateur.owner.full_name if indicateur.owner else "",
            ])
            total += 1

        _ajuster(feuille)

        # ---- historique des releves valides
        feuille2 = classeur.create_sheet("Releves")
        _entete(feuille2, [
            "Indicateur", "Debut de periode", "Fin de periode",
            "Valeur atteinte", "Statut", "Commentaire",
        ])
        for indicateur in indicateurs:
            for releve in indicateur.readings.all():
                feuille2.append([
                    indicateur.code,
                    releve.period_start,
                    releve.period_end,
                    float(releve.achieved_value),
                    releve.get_status_display(),
                    releve.comment,
                ])
                total += 1
        _ajuster(feuille2)

        return total